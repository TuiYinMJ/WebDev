import os
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.exceptions import InvalidFileException
from pathlib import Path

excel_file = 'find.xlsx'
sheet_name = 'Sheet1'

root_dir = r'D:\图片'
save_dir = r'D:\FILE_FIND'

# 初始化一个全局的唯一文件名计数器
unique_file_counter = 0

try:
    workbook = load_workbook(excel_file)
    sheet = workbook[sheet_name]

    def search_and_copy(item_number, quantity, group_dir):
        global unique_file_counter
        print(f"Searching for {item_number} in {group_dir}...")
        item_number_lower = item_number.lower()
        files_to_copy = [
            (root, file) for root, _, files in os.walk(root_dir) for file in files
            if item_number_lower in file.lower()
        ]
        copied_files = 0
        for root, file in files_to_copy:
            src_path = Path(root) / file
            for _ in range(quantity):
                # 使用全局唯一计数器来确保文件名的唯一性
                dst_filename = f"{item_number}_{unique_file_counter}{os.path.splitext(file)[1]}"
                unique_file_counter += 1  # 每次复制后递增计数器
                dst_path = group_dir / dst_filename
                shutil.copy2(src_path, dst_path)
                copied_files += 1
        return copied_files

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        group, item_number, quantity = row
        if group is None or not group.strip():
            continue

        group_dir = Path(save_dir) / group
        group_dir.mkdir(parents=True, exist_ok=True)

        copied_files = search_and_copy(item_number, quantity, group_dir)
        if copied_files < quantity:
            yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
            sheet.cell(row=idx, column=3).fill = yellow_fill

    workbook.save('findx.xlsx')
except InvalidFileException as e:
    print(f"Error: {e}")
except PermissionError as e:
    print(f"Error: {e.strerror}")
except shutil.SameFileError as e:
    print(f"Error: {e.strerror}")
except OSError as e:
    print(f"Error: {e.strerror}")
except InvalidFileException as e:
    print(f"Error: {e}")
except Exception as e:
    print(f"Unexpected error: {e}")
finally:
    try:
        # 即使遇到错误，也要尝试保存Excel文件
        workbook.save('findx.xlsx')
    except Exception as e:
        print(f"Failed to save Excel file: {e}")
