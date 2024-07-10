import os
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.exceptions import InvalidFileException
from pathlib import Path
# import subprocess

excel_file = 'find.xlsx'
sheet_name = 'Sheet1'


root_dir = r'D:\图片'   # 查找文件夹
save_dir = r'D:\FILE_FIND\gg'  # 保存路径



try:
    # 打开Excel文件
    wordb = load_workbook(excel_file)
    sheet = wordb[sheet_name]

    # 定义一个函数来递归搜索文件
    def search_and_copy(item_number, quantity, group_dir):
        print(f"Searching for {item_number} in {group_dir}...")
        item_number_lower = item_number.lower()
        files_to_copy = [
            (root, file) for root, _, files in os.walk(root_dir) for file in files
            if item_number_lower in file.lower()
        ]
        copied_files = 0
        for root, file in files_to_copy:
            src_path = Path(root) / file
            for i in range(quantity):
                # 修改文件名以避免覆盖
                dst_filename = f"{item_number}_{i}{file[-4:]}"
                dst_path = group_dir / dst_filename
                shutil.copy2(src_path, dst_path)
                # subprocess.run(['cmd', '/c', 'copy', '/y', str(src_path), str(dst_path)], check=True)
                copied_files += 1
        return copied_files


    # 创建文件夹和复制/标记文件
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        group, item_number, quantity = row
        if group is None or not group.strip():
            continue

        group_dir = Path(save_dir) / group
        group_dir.mkdir(parents=True, exist_ok=True)

        # 检查并复制文件
        copied_files = search_and_copy(item_number, quantity, group_dir)
        if copied_files < quantity:
            # 如果复制的文件数量少于要求的数量，设置数量单元格的背景颜色为黄色
            yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
            sheet.cell(row=idx, column=2).fill = yellow_fill

    # 保存Excel文件
    wordb.save('findx.xlsx')
except FileNotFoundError as e:
    print(f"Error: {e.strerror}")
except PermissionError as e:
    print(f"Error: {e.strerror}")
except shutil.SameFileError as e:
    print(f"Error: {e.strerror}")
except OSError as e:
    print(f"Error: {e.strerror}")
except InvalidFileException as e:
    print(f"Error: {e}")
except Exception as e:
    print(f"Unexpected error: {e}")
finally:
    try:
        # 即使遇到错误，也要尝试保存Excel文件
        wordb.save('findx.xlsx')
    except Exception as e:
        print(f"Failed to save Excel file: {e}")
